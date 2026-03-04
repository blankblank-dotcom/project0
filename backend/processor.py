"""
backend/processor.py — The Local Titan: Document Processor
==========================================================

Handles all pre-inference document preparation:
  1. PDF → PIL.Image conversion (multi-page, via pypdf)
  2. Dynamic resizing for optimal OCR accuracy vs CPU latency
  3. Auto-redaction of PII regions using localized Gaussian blur

Design constraints:
  - Must work on 16GB RAM (6GB OS + UI headroom)
  - Pages are processed one-at-a-time (streaming, not batch-loaded)
  - Target image width: 1024px–1280px (sweet spot for VLM accuracy)
"""

from __future__ import annotations

import io
import logging
import asyncio
from dataclasses import dataclass
from pathlib import Path
from typing import Generator, Optional, AsyncGenerator

from PIL import Image, ImageFilter, ImageDraw

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
log = logging.getLogger("processor")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
# Dynamic resize bounds (pixels)
TARGET_WIDTH_MIN = 1024
TARGET_WIDTH_MAX = 1920
TARGET_WIDTH_DEFAULT = 1600

# PDF rendering DPI — controls quality of rasterization
# 150 DPI ≈ 1240px width for US Letter → falls in our sweet spot
# 200 DPI ≈ 1654px width → we'd downscale, but get crisper text
RENDER_DPI = 150

# Auto-redact blur radius
REDACT_BLUR_RADIUS = 30  # Gaussian kernel radius for PII redaction


# ---------------------------------------------------------------------------
# Data Structures
# ---------------------------------------------------------------------------
@dataclass
class PageImage:
    """A single rasterized page with metadata.

    Attributes:
        image:        PIL Image of the page (RGB, resized).
        page_number:  1-indexed page number.
        original_size: (width, height) before resizing.
        source_file:  Path to the source PDF.
    """
    image: Image.Image
    page_number: int
    original_size: tuple[int, int]
    source_file: str


# ═══════════════════════════════════════════════════════════════════════════
# CONCURRENCY CONTROL (P-core Protection)
# ═══════════════════════════════════════════════════════════════════════════

# Global Semaphore to limit active OpenVINO inference to exactly 1.
# This ensures that a single inference stream has exclusive access to
# the CPU's massive L3 cache and Performance (P) cores, maximizing
# throughput on hybrid Intel architectures.
INFERENCE_SEMAPHORE = asyncio.Semaphore(1)


class ProcessingQueue:
    """An asynchronous queue for staging document pages for inference.
    
    This allows PDF-to-image extraction (on E-cores) to run ahead of
    AI inference (on P-cores), ensuring the pipeline is never stalled.
    
    Features:
      - Max capacity to prevent RAM exhaustion.
      - Pause/Resume mechanism for system health (RAM safety).
    """
    def __init__(self, maxsize: int = 10):
        self._queue = asyncio.Queue(maxsize=maxsize)
        self._is_paused = False
        self._pause_event = asyncio.Event()
        self._pause_event.set() # Default = Not paused

    async def put(self, page: PageImage):
        """Add a page to the queue. Blocks if queue is full."""
        await self._queue.put(page)

    async def get(self) -> PageImage:
        """Retrieve a page from the queue. Blocks if paused or empty."""
        await self._pause_event.wait()
        return await self._queue.get()

    def task_done(self):
        """Signal that a retrieved page has been processed."""
        self._queue.task_done()

    @property
    def qsize(self) -> int:
        return self._queue.qsize()

    def pause(self):
        """Halt extraction — used when system RAM is critical."""
        if not self._is_paused:
            log.warning("ProcessingQueue PAUSED due to system health.")
            self._is_paused = True
            self._pause_event.clear()

    def resume(self):
        """Resume extraction."""
        if self._is_paused:
            log.info("ProcessingQueue RESUMED.")
            self._is_paused = False
            self._pause_event.set()

    @property
    def is_paused(self) -> bool:
        return self._is_paused


# ═══════════════════════════════════════════════════════════════════════════
# 1. PDF → PIL.Image EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════
def extract_pages_from_pdf(
    pdf_path: str | Path,
    dpi: int = RENDER_DPI,
    max_pages: Optional[int] = None,
) -> Generator[PageImage, None, None]:
    """Extract pages from a PDF as PIL Images.

    Uses pypdf to extract embedded images or, as a fallback, renders
    pages via pdf2image (poppler). For maximum compatibility on Windows,
    the primary path uses pypdf's built-in image extraction.

    Args:
        pdf_path:   Path to the PDF file.
        dpi:        Render resolution (affects image width).
        max_pages:  Maximum pages to extract (None = all).

    Yields:
        PageImage objects, one per page, in order.
    """
    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path.resolve()}")

    if not pdf_path.suffix.lower() == ".pdf":
        raise ValueError(f"Expected a .pdf file, got: {pdf_path.suffix}")

    log.info(f"Opening PDF: {pdf_path.name}")

    try:
        # Try the poppler-based approach first (best quality)
        yield from _extract_via_pdf2image(pdf_path, dpi, max_pages)
    except (ImportError, Exception) as e:
        log.warning(
            f"pdf2image unavailable ({e}), falling back to pypdf extraction."
        )
        yield from _extract_via_pypdf(pdf_path, max_pages)


def _extract_via_pdf2image(
    pdf_path: Path,
    dpi: int,
    max_pages: Optional[int],
) -> Generator[PageImage, None, None]:
    """Render PDF pages to images using pdf2image (poppler backend)."""
    from pdf2image import convert_from_path

    pages = convert_from_path(
        str(pdf_path),
        dpi=dpi,
        first_page=1,
        last_page=max_pages,
        fmt="RGB",
        thread_count=2,  # Limit threads for memory safety
    )

    for i, page_img in enumerate(pages, start=1):
        original_size = page_img.size
        resized = dynamic_resize(page_img)

        log.info(
            f"  Page {i}: {original_size[0]}×{original_size[1]} → "
            f"{resized.size[0]}×{resized.size[1]}"
        )

        yield PageImage(
            image=resized,
            page_number=i,
            original_size=original_size,
            source_file=str(pdf_path),
        )

        # Explicitly free the original if we made a copy
        if resized is not page_img:
            del page_img


def _extract_via_pypdf(
    pdf_path: Path,
    max_pages: Optional[int],
) -> Generator[PageImage, None, None]:
    """Extract embedded images from PDF pages using pypdf.

    This fallback works without poppler but may produce lower quality
    results for scanned documents (only extracts embedded image objects,
    does not render vector content).
    """
    from pypdf import PdfReader

    reader = PdfReader(str(pdf_path))
    total = len(reader.pages)
    limit = min(total, max_pages) if max_pages else total

    log.info(f"  PDF has {total} page(s), processing {limit}.")

    for page_idx in range(limit):
        page = reader.pages[page_idx]
        page_num = page_idx + 1

        # Try to extract the largest embedded image from the page
        images = page.images
        if images:
            # Take the largest image (by data size) as the page image
            largest = max(images, key=lambda img: len(img.data))
            pil_img = Image.open(io.BytesIO(largest.data)).convert("RGB")
        else:
            # No embedded images — create a placeholder
            log.warning(
                f"  Page {page_num}: No embedded images found. "
                f"This page may be vector-only (needs poppler)."
            )
            # Create a white placeholder so the pipeline doesn't break
            pil_img = Image.new("RGB", (TARGET_WIDTH_DEFAULT, 1600), "white")

        original_size = pil_img.size
        resized = dynamic_resize(pil_img)

        log.info(
            f"  Page {page_num}: {original_size[0]}×{original_size[1]} → "
            f"{resized.size[0]}×{resized.size[1]}"
        )

        yield PageImage(
            image=resized,
            page_number=page_num,
            original_size=original_size,
            source_file=str(pdf_path),
        )


# ═══════════════════════════════════════════════════════════════════════════
# 2. DYNAMIC RESIZER
# ═══════════════════════════════════════════════════════════════════════════
def dynamic_resize(
    image: Image.Image,
    min_width: int = TARGET_WIDTH_MIN,
    max_width: int = TARGET_WIDTH_MAX,
) -> Image.Image:
    """Resize an image to the optimal width for VLM inference.

    Strategy:
      - If width is already within [min_width, max_width], return as-is.
      - If wider than max_width, downscale to max_width (preserving aspect).
      - If narrower than min_width, upscale to min_width (preserving aspect).

    The goal is to balance OCR/VLM accuracy (needs sufficient resolution)
    with CPU latency (larger images = more vision tokens = slower).

    Args:
        image:     Input PIL Image.
        min_width: Minimum acceptable width (default 1024px).
        max_width: Maximum acceptable width (default 1280px).

    Returns:
        Resized PIL Image (or the original if already in range).
    """
    w, h = image.size

    if min_width <= w <= max_width:
        # Already in the sweet spot
        return image

    if w > max_width:
        target_w = max_width
    else:
        target_w = min_width

    # Preserve aspect ratio
    scale = target_w / w
    target_h = int(h * scale)

    resized = image.resize(
        (target_w, target_h),
        Image.Resampling.LANCZOS,
    )

    log.debug(f"Resized: {w}×{h} → {target_w}×{target_h} (scale={scale:.2f})")
    return strip_metadata(resized)


def strip_metadata(image: Image.Image) -> Image.Image:
    """Remove all EXIF, ICC profiles, and auxiliary metadata from a PIL image.
    
    This ensures that location data, camera info, and source software
    identity are purged before the document is processed or exported.
    """
    dataSize = len(image.tobytes())
    # Create a fresh image copy without the .info dict
    clean = Image.new(image.mode, image.size)
    clean.putdata(image.getdata())
    log.info(f"  Metadata purged (Scrubbed {dataSize} bytes of pixel mapping) ✓")
    return clean


# ═══════════════════════════════════════════════════════════════════════════
# 3. AUTO-REDACT (PII Gaussian Blur)
# ═══════════════════════════════════════════════════════════════════════════
def auto_redact(
    image: Image.Image,
    regions: list[dict],
    blur_radius: int = REDACT_BLUR_RADIUS,
) -> Image.Image:
    """Apply localized Gaussian blur to PII regions on a document image.

    Only the pixels within each specified bounding box are blurred;
    the rest of the document remains untouched.

    Args:
        image:       PIL Image to redact (will NOT be mutated — returns a copy).
        regions:     List of region dicts, each with:
                       - "label": str (e.g. "SSN", "Full Name")
                       - "box_2d": [ymin, xmin, ymax, xmax] normalized [0.0, 1.0]
                     This matches the output format from engine.GroundedObject.
        blur_radius: Gaussian blur kernel radius. Higher = more obscured.

    Returns:
        A new PIL Image with PII regions blurred.
    """
    if not regions:
        log.info("No regions to redact.")
        return strip_metadata(image)

    # Work on a copy to avoid mutating the original
    redacted = image.copy()
    draw = ImageDraw.Draw(redacted)
    img_w, img_h = redacted.size

    redacted_count = 0

    for region in regions:
        label = region.get("label", "unknown")
        box = region.get("box_2d", [])

        if len(box) != 4:
            log.warning(f"Skipping region '{label}': invalid box_2d={box}")
            continue

        ymin, xmin, ymax, xmax = box

        # Convert normalized coords to pixel coords
        px_left = int(xmin * img_w)
        px_top = int(ymin * img_h)
        px_right = int(xmax * img_w)
        px_bottom = int(ymax * img_h)

        # Clamp to image bounds
        px_left = max(0, min(img_w, px_left))
        px_top = max(0, min(img_h, px_top))
        px_right = max(0, min(img_w, px_right))
        px_bottom = max(0, min(img_h, px_bottom))

        # Skip degenerate regions
        if px_right <= px_left or px_bottom <= px_top:
            continue

        # Burn-in: Physically replace pixels with solid black
        draw.rectangle([px_left, px_top, px_right, px_bottom], fill="black", outline="black")

        redacted_count += 1
        log.info(f"  [BURNED] Redacted '{label}': ({px_left},{px_top})→({px_right},{px_bottom})")

    log.info(f"✓ DLP Redaction complete: {redacted_count}/{len(regions)} regions physically purged.")
    return strip_metadata(redacted)


def auto_redact_from_engine_result(
    image: Image.Image,
    engine_result,
    blur_radius: int = REDACT_BLUR_RADIUS,
) -> Image.Image:
    """Convenience wrapper that accepts an EngineResult directly.

    Extracts the grounded objects from the engine result and redacts them.

    Args:
        image:         PIL Image to redact.
        engine_result: An EngineResult from engine.InferenceEngine.detect_pii().
        blur_radius:   Gaussian blur kernel radius.

    Returns:
        A new PIL Image with PII regions blurred.
    """
    regions = [obj.to_dict() for obj in engine_result.objects]
    return auto_redact(image, regions, blur_radius)


# ═══════════════════════════════════════════════════════════════════════════
# CONVENIENCE: Full Pipeline Helper
# ═══════════════════════════════════════════════════════════════════════════
def load_single_image(
    path: str | Path,
    resize: bool = True,
) -> Image.Image:
    """Load a single image file (PNG, JPG, TIFF) for inference.

    Args:
        path:   Path to the image file.
        resize: Whether to apply dynamic resizing.

    Returns:
        PIL Image ready for the inference engine.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Image not found: {path.resolve()}")

    img = Image.open(path).convert("RGB")
    log.info(f"Loaded image: {path.name} ({img.size[0]}×{img.size[1]})")

    if resize:
        img = dynamic_resize(img)

    return img


# ═══════════════════════════════════════════════════════════════════════════
# 6. MEMORY CONVERSION (PIL → NumPy / Raw)
# ═══════════════════════════════════════════════════════════════════════════
def pil_to_numpy_raw(image: Image.Image) -> "np.ndarray":
    """Convert a PIL Image to a raw NumPy array (HWC, RGB).
    
    Used for the OpenVINO Preprocessing API to bypass PIL's internal
    resizing logic and use iGPU/E-core hardware instead.
    """
    import numpy as np
    
    if image.mode != "RGB":
        image = image.convert("RGB")
        
    return np.array(image, dtype=np.uint8)



# ═══════════════════════════════════════════════════════════════════════════
# 5. DATA EXPORTER (JSON → Excel / CSV)
# ═══════════════════════════════════════════════════════════════════════════
class DataExporter:
    """Convert VLMPipeline extraction results to exportable file formats.

    Supported exports:
      - .xlsx  (pandas + openpyxl, with styled header)
      - .csv   (pandas fallback if openpyxl not available)

    Usage:
        exporter = DataExporter(output_data, source_filename="invoice.pdf")
        path = exporter.to_excel()       # → /tmp/LocalTitan/invoice_results_20260303.xlsx
        path = exporter.to_csv()         # → /tmp/LocalTitan/invoice_results_20260303.csv
    """

    def __init__(
        self,
        output_data: list[dict[str, str]],
        source_filename: str = "document",
        bounding_boxes: Optional[list[dict]] = None,
    ):
        self._data = output_data
        self._source = Path(source_filename).stem
        self._boxes = bounding_boxes or []

    def _get_export_dir(self) -> Path:
        """Writable export directory (AppData on Windows, /tmp otherwise)."""
        import sys
        import os
        import tempfile

        if sys.platform == "win32":
            base = Path(os.environ.get("LOCALAPPDATA", tempfile.gettempdir()))
            export_dir = base / "LocalTitan" / "exports"
        else:
            export_dir = Path(tempfile.gettempdir()) / "LocalTitan" / "exports"

        export_dir.mkdir(parents=True, exist_ok=True)
        return export_dir

    def _make_filename(self, ext: str) -> Path:
        """Generate a timestamped filename."""
        from datetime import datetime

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in self._source)
        return self._get_export_dir() / f"{safe_name}_results_{stamp}.{ext}"

    def to_excel(self) -> Path:
        """Export to a styled .xlsx workbook.

        Returns:
            Path to the generated Excel file.
        """
        import pandas as pd

        df = pd.DataFrame(self._data)
        output_path = self._make_filename("xlsx")

        try:
            from openpyxl.styles import Font, PatternFill, Alignment

            with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
                # Data sheet
                df.to_excel(writer, sheet_name="Extracted Data", index=False)
                ws = writer.sheets["Extracted Data"]

                # Style header row
                header_fill = PatternFill(
                    start_color="18181b", end_color="18181b", fill_type="solid"
                )
                header_font = Font(name="Segoe UI", bold=True, color="fafafa", size=11)

                for col_idx, cell in enumerate(ws[1], 1):
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left")

                # Auto-fit column widths
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

                # Bounding boxes sheet (if any)
                if self._boxes:
                    box_df = pd.DataFrame(self._boxes)
                    box_df.to_excel(writer, sheet_name="Bounding Boxes", index=False)

                # Metadata sheet
                meta_df = pd.DataFrame([
                    {"Property": "Source File", "Value": self._source},
                    {"Property": "Total Fields", "Value": str(len(self._data))},
                    {"Property": "Export Tool", "Value": "The Local Titan"},
                ])
                meta_df.to_excel(writer, sheet_name="Metadata", index=False)

        except ImportError:
            log.warning("openpyxl not available — writing plain xlsx via pandas.")
            df.to_excel(str(output_path), index=False)

        log.info(f"Exported Excel: {output_path}")
        return output_path

    def to_csv(self) -> Path:
        """Export to a plain CSV file.

        Returns:
            Path to the generated CSV file.
        """
        import pandas as pd

        df = pd.DataFrame(self._data)
        output_path = self._make_filename("csv")
        df.to_csv(str(output_path), index=False)

        log.info(f"Exported CSV: {output_path}")
        return output_path


# ═══════════════════════════════════════════════════════════════════════════
# 6. BATCH RESULT PERSISTENCE (SQLite)
# ═══════════════════════════════════════════════════════════════════════════
from backend.db import _security
import sqlite3 as _sqlite3
import json as _json


def _get_results_db_path() -> Path:
    """Path to the SQLite database for persisting batch results."""
    import sys
    import os
    import tempfile

    if sys.platform == "win32":
        base = Path(os.environ.get("LOCALAPPDATA", tempfile.gettempdir()))
        db_dir = base / "LocalTitan"
    else:
        db_dir = Path(tempfile.gettempdir()) / "LocalTitan"

    db_dir.mkdir(parents=True, exist_ok=True)
    return db_dir / "batch_results.db"


def _ensure_results_table(conn: _sqlite3.Connection) -> None:
    """Create the batch results table if it doesn't exist."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS batch_results (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            filename    TEXT NOT NULL,
            page_num    INTEGER DEFAULT 1,
            field_key   TEXT,
            field_value TEXT,
            raw_output  TEXT,
            processed_at TEXT DEFAULT (datetime('now'))
        )
        """
    )
    conn.commit()


def save_batch_result(
    filename: str,
    page_num: int,
    output_data: list[dict[str, str]],
    raw_output: str = "",
) -> None:
    """Persist a single page's extraction results to SQLite immediately.

    This is called after EACH page in a batch to ensure no data loss
    if the app is closed mid-batch.

    Args:
        filename:    Source file name.
        page_num:    1-indexed page number.
        output_data: List of {\"field\": ..., \"value\": ...} dicts.
        raw_output:  Raw model output string.
    """
    db_path = _get_results_db_path()
    conn = _sqlite3.connect(str(db_path))
    try:
        _ensure_results_table(conn)
        for row in output_data:
            field_key = row.get("field", "")
            field_val = row.get("value", "")
            
            # Encrypt sensitive values
            enc_val = _security.encrypt(field_val)
            enc_raw = _security.encrypt(raw_output)
            
            conn.execute(
                "INSERT INTO batch_results (filename, page_num, field_key, field_value, raw_output) "
                "VALUES (?, ?, ?, ?, ?)",
                (filename, page_num, field_key, enc_val, enc_raw),
            )
        conn.commit()
        log.info(f"Saved {len(output_data)} result(s) for '{filename}' page {page_num}.")
    finally:
        conn.close()


def export_batch_to_excel(batch_filenames: Optional[list[str]] = None) -> Path:
    """Export all batch results from SQLite to a single Excel file.

    Args:
        batch_filenames: Optional filter — only export these files. None = all.

    Returns:
        Path to the exported Excel file.
    """
    import pandas as pd

    db_path = _get_results_db_path()
    conn = _sqlite3.connect(str(db_path))

    try:
        if batch_filenames:
            placeholders = ", ".join("?" * len(batch_filenames))
            query = f"SELECT * FROM batch_results WHERE filename IN ({placeholders})"
            df = pd.read_sql_query(query, conn, params=batch_filenames)
        else:
            df = pd.read_sql_query("SELECT * FROM batch_results", conn)
    finally:
        conn.close()

    exporter = DataExporter(
        df[["field_key", "field_value"]].rename(
            columns={"field_key": "field", "field_value": "value"}
        ).to_dict("records"),
        source_filename="batch_export",
    )
    return exporter.to_excel()

